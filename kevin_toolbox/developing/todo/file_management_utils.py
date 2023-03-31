# 源文件位置：[windows laptop]\Desktop\神经网络学习\常用代码\file_management_utils.py
# 整理至：data_flow.file
import json
import os
import re
import time
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

'''
path

检查:
convert_path_to_standard_form()
    把任意路径 path 转换成当前系统的格式

生成相对路径:
get_relative_path()
    获取相对路径

程序根路径:
root_path
'''

root_path = os.path.abspath(os.path.dirname(__file__))


def convert_path_to_standard_form(path):
    """
    把任意路径 path 转换成当前系统的格式
    """

    seps = r'\/'
    sep_other = seps.replace(os.sep, '')
    path = path.replace(sep_other, os.sep)
    while os.sep * 2 in path:
        path.replace(os.sep * 2, os.sep)
    return path


def get_relative_path(base_path, long_path):
    """
    获取相对路径

    参数:
        base_path: string 文件夹路径。
        long_path: string 文件/子文件夹路径。

    返回:
        relative_path 文件/子文件夹相对于文件夹的路径，为 a/b/c 形式
    """

    if base_path in long_path:
        relative_path = long_path.replace(base_path, '')
        relative_path = relative_path[1:] if relative_path[0] == os.sep else relative_path
        return relative_path
    else:
        return long_path


'''
file/folder name

生成:
assemble_name()
    生成文件/文件夹名
    返回  name : " base(__dtime__date)(.suffix) "

解释:
disassemble_name()
    分解文件/文件夹名
    返回  name_component_dict : {'base': base, 'dtime': dtime, 'date': date, 'suffix': suffix}

检查: 
modify_name()
    去除 文件/文件夹名 中的非法字符

重命名:
rename()
    重命名 pre_path/old_name  --rename-->  pre_path/new_name


常用参数/变量:
base: 名字主体
suffix: 后缀（默认带.）
dtime: 绝对时间戳，由time.time_ns()获取
date: 日期

pre_path: 路径中的前段  pre_path, folder/file_name = os.path.split(path)
'''

dtime_of_last_assemble_name = None


def assemble_name(base, suffix=None, use_timestamp=True):
    """
    生成文件/文件夹名  " base(__dtime__date)(.suffix) "

    参数:
        base: string 名字主体。
        suffix: string or None 文件后缀。如果不设定，则生成文件夹名。
        use_timestamp: boolean 是否加上时间戳。

    返回:
        string 文件名
        " name(__dtime__date)(.suffix) "
        注：括号内为可选部分
    """

    if base is None:
        return None

    if use_timestamp:
        dtime = time.time_ns()
        global dtime_of_last_assemble_name
        while dtime == dtime_of_last_assemble_name:  # 避免重名
            time.sleep(1e-10)
            dtime = time.time_ns()
        dtime_of_last_assemble_name = dtime

        date = time.strftime("%Y-%m-%d,%H-%M-%S", time.localtime(dtime * 1e-9))
        base = '__'.join([base, str(dtime), date])

    if suffix:
        if suffix.startswith('.'):
            suffix = suffix.split('.', 1)[1]
        name = '.'.join([base, suffix])
    else:
        name = base

    return name


rule_base_2_base_dtime_date = re.compile(r"(?P<base>.+?)__(?P<dtime>.+?)__(?P<date>.+)")


def disassemble_name(name, suffix_exist=True):
    """
    分解文件/文件夹名  " base(__dtime__date)(.suffix) "

    参数:
        name: string
        suffix_exist: boolean 是否存在/需要分解后缀名

    返回:
        name_component_dict : dict {'base': base, 'dtime': dtime, 'date': date, 'suffix': suffix}
    """

    if name is None:
        return None

    global rule_base_2_base_dtime_date

    # （如果想要分解）后缀（则更新base和suffix）
    if suffix_exist:
        base, suffix = os.path.splitext(name)
    else:
        base = name
        suffix = None

    # （先默认没有）时间戳
    dtime, date = None, None
    # （如果找到）时间戳（则更新dtime和date）
    search_base = rule_base_2_base_dtime_date.search(base)
    if search_base:  # 带有时间戳  base__dtime__date
        base = search_base.group('base')
        dtime = int(search_base.group('dtime'))
        date = search_base.group('date')

    name_component_dict = {'base': base, 'dtime': dtime, 'date': date, 'suffix': suffix}

    return name_component_dict


def rename(pre_path, old_name, new_name, conflict_strategy='skip'):
    """
    重命名  pre_path/old_name  --rename-->  pre_path/new_name

    参数:
        conflict_strategy: 冲突处理方式  one in {'preserve_old', 'preserve_new', 'skip'}
            preserve_old: 保留 old_name 指向的文件
            preserve_new: 保留 new_name 指向的文件
            skip: 不作处理

    返回:
        boolean 是否成功重命名
    """

    assert conflict_strategy in {'preserve_old', 'preserve_new', 'skip'}

    path_old = os.path.join(pre_path, old_name)
    path_new = os.path.join(pre_path, new_name)

    # 旧文件（不存在则退出）
    if not os.path.exists(path_old):
        print("Warn: source_path does not exist.")
        return False

    # 目标文件（若已存在，则进行冲突处理）
    if os.path.exists(path_new):
        print("Warn: target path already exist.\n      Conduct conflict strategy: " + conflict_strategy)
        if conflict_strategy == 'preserve_old':
            os.remove(path_new)
            os.rename(path_old, path_new)
        elif conflict_strategy == 'preserve_new':
            os.remove(path_old)
        elif conflict_strategy == 'skip':
            return False
    else:
        os.rename(path_old, path_new)

    return True


def modify_name(name):
    """
    去除 文件/文件夹名 中的非法字符
    文件/文件夹名不能含有\\ /:*?"<>|
    """

    for char in r'\/:*?"<>|':
        name = name.replace(char, "", -1)
    return name


'''
manage files and folders 文件管理

新建:
make_folder()
    新建文件夹
make_ls_of_folder()
    新建一系列文件夹

移除:
remove()
    移除文件/文件夹

查找文件: 
search_file()
    查找 folder 目录中，带指定后缀 suffix 和 flag 的文件的路径
    分级返回结果
    各级 folder_path 及其下面符合条件的文件名 [{"folder_path": folder_path, "file_name_ls": file_name_ls}, ……]

search_file_flat()
    结果不分级  list of file_path（绝对路径）
'''


def make_folder(*path):
    """
    新建文件夹

    返回:
        already_existed, path
        already_existed
            True: 文件夹已经存在
            False: 不存在，新建成功
    """

    path_com = os.path.join(*path)

    if not os.path.isdir(path_com):  # 不存在则建立
        os.makedirs(path_com)
        return False, path_com
    else:
        return True, path_com


def make_ls_of_folder(*path_ls):
    """
    新建一系列文件夹

    返回:
        success_path: list of path 新建的文件夹的路径
    """

    success_path = []
    for path in path_ls:
        already_existed, path = make_folder(path)
        if not already_existed:
            success_path.append(path)
    return success_path


def remove(*path):
    """
    移除文件/文件夹

    返回:
        boolean 是否成功
    """

    path_com = os.path.join(*path)

    try:
        if os.path.isfile(path_com):  # 移除文件
            os.remove(path_com)
            # print("文件%s已移除" % path_com)
        elif os.path.isdir(path_com):  # 移除文件夹
            os.removedirs(path_com)
            # print("文件夹%s已移除" % path_com)
        return True
    except Exception as e:  # 删除失败
        print("Warn: fail to delete.\n      path: " + path_com)
        print(Exception, ":", e)
        return False


def search_file(folder, suffix_ls=None, flag_ls=None, case_sensitive=False):
    """
    search all file_path_dict_ls with specific suffix in folder (Hierarchical)

    获取 folder 目录中，带指定后缀 suffix 和 flag 的文件的路径
    suffix_ls 和 flag_ls 不指定时表示不开启筛选，直接返回 folder 目录下的所有文件路径

    参数:
        folder: directory path to be searched  要求是绝对路径
        suffix_ls: list of suffix
        flag_ls: list of flag , flag is the tag in the file name
        case_sensitive: boolean 是否区分大小写（默认为False不区分）

    返回:
        folder_file_dict_ls
        各级 folder_path 及其下面符合条件的文件名 [{"folder_path": folder_path, "file_name_ls": file_name_ls}, ……]
    """

    dir_ls = os.listdir(folder)  # 获取path目录下的文件和文件夹
    file_name_ls = []
    folder_file_dict_ls = []
    for dir_ in dir_ls:
        path_tmp = os.path.join(folder, dir_)
        if os.path.isdir(path_tmp):  # 如是目录,则递归查找
            folder_file_dict_ls += search_file(path_tmp, suffix_ls, flag_ls)
        else:  # 不是目录,则比较后缀名
            file_name_ls.append(dir_)  # 先加进来，后面再剔除

            if not case_sensitive:
                dir_ = dir_.lower()

            name, suffix = os.path.splitext(dir_)
            if suffix_ls and suffix not in suffix_ls:  # suffix 筛选
                file_name_ls.pop(-1)
            elif flag_ls:  # flag 筛选
                bingo = False
                for flag in flag_ls:
                    if flag in name:
                        bingo = True
                if not bingo:
                    file_name_ls.pop(-1)

    folder_file_dict_ls = [{"folder_path": folder, "file_name_ls": file_name_ls}] + folder_file_dict_ls
    return folder_file_dict_ls


def search_file_flat(folder, suffix_ls=None, flag_ls=None, case_sensitive=False):
    """
    search all file_path_dict with specific suffix in folder (Regardless of the hierarchy)

    输入参数介绍请参见 search_file() 方法
    """

    file_path_ls = []
    folder_file_dict_ls = search_file(folder, suffix_ls, flag_ls, case_sensitive)
    for folder_filename_dict in folder_file_dict_ls:
        _folder_ = folder_filename_dict.get("folder_path")
        for _filename_ in folder_filename_dict.get("file_name_ls"):
            file_path = os.path.join(_folder_, _filename_)
            file_path_ls.append(file_path)
    return file_path_ls


'''
read and write

markdown:
read_markdown()
write_markdown()

txt:
read_txt()  不是通用的，需要根据实际修改读取格式

json: 
read_json()
write_json()

json: 
read_npy()
write_npy()


search_file_flat()
    结果不分级  list of file_path（绝对路径）
'''


# markdown
def read_markdown(*path):
    try:
        path_com = os.path.join(*path)
        if os.path.isfile(path_com):
            with open(path_com, 'r', encoding='utf-8') as file:
                content = file.read()
            print(path_com, ' was read.')
            return content
        else:
            print("Warn: file does not exist.\n      path: " + path_com)
            return None
    except Exception as e:
        print("Error: failed to read.")
        print(Exception, ":", e)
        return None


def write_markdown(content, *path):
    if content:
        try:
            path_com = os.path.join(*path)
            with open(path_com, 'w', encoding='utf-8') as fff:
                fff.write(content)
            print(path_com, 'wrote.')
            return path_com
        except Exception as e:
            print("failed to write")
            print(Exception, ":", e)
            return None
    else:
        return None


# txt
def read_txt(*path):
    """
    按行读取 txt 中的 url
    """

    path_com = os.path.join(*path)
    if os.path.isfile(path_com):
        with open(path_com, 'r') as file:
            content = file.read()
            row_ls = content.split('\n')
        url_ls_res = []
        for url_ in row_ls:
            url_ = url_.replace(' ', '')  # 删除空格
            url_ls_res.append(url_)
        url_ls_res = list(set(url_ls_res))
        return url_ls_res
    else:
        print("文件名错误！")
        return False


# numpy
def read_npy(*path):
    path_com = os.path.join(*path)
    try:
        res = np.load(path_com)
    except:  # 发生异常，执行这块代码
        print('Error: failed to read.\n       path: ' + path_com)
        return np.array([])
    else:  # 如果没有异常执行这块代码
        return res


def write_npy(content, *path):
    path_com = os.path.join(*path)
    try:
        np.save(path_com, content)
    except:  # 发生异常，执行这块代码
        print('Error: failed to write.\n       path: ' + path_com)
        print('失败内容: ', content)
        return False
    else:  # 如果没有异常执行这块代码
        return True


# excel
def write_excel_with_matrix(file_path_abs, sheet_name, matrix, column_label_ls=None, row_label_ls=None, column_title="",
                            row_title="", main_title=""):
    """
    将矩阵写入到 excel 文件中

    获取 folder 目录中，带指定后缀 suffix 和 flag 的文件的路径
    suffix_ls 和 flag_ls 不指定时表示不开启筛选，直接返回 folder 目录下的所有文件路径

    参数:
        file_path_abs: 文件绝对路径
        matrix: 矩阵  np.array or np.matrix
        column_label_ls, row_label_ls: 行列标签
        column_title, row_title: 行列标题
        main_title: 总标题
    """

    # 判断文件是否存在，不存在则新建，否则读取文件
    if not os.path.isfile(file_path_abs):
        wb = openpyxl.Workbook()  # 创建文件对象
        # wb对象创建后，默认含有一个默认的名为 Sheet 的 页面,将其删除
        ws_ = wb.active
        wb.remove(ws_)
    else:
        wb = openpyxl.load_workbook(file_path_abs)
    # 判断sheet是否存在，不存在则建立，否则先删除再建立
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    # print(sheet_name)
    ws = wb.create_sheet(sheet_name)

    # 开始写
    matrix_r_offset, matrix_c_offset = 1, 1  # 矩阵的起始位置
    r_offset, c_offset = 1, 1  # 目前的写入位置
    for i in [main_title, column_title, column_label_ls]:
        if i:
            matrix_r_offset += 1
    for j in [row_title, row_label_ls]:
        if j:
            matrix_c_offset += 1
    # print(matrix)
    matrix_row_num, matrix_column_num = matrix.shape[0], matrix.shape[1]
    # 标题
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if main_title:
        ws.merge_cells(start_row=r_offset, start_column=1, end_row=r_offset,
                       end_column=matrix_c_offset + matrix_column_num - 1)
        ws.cell(row=r_offset, column=1).value = main_title
        ws.cell(row=r_offset, column=1).alignment = alignment
        ws.cell(row=r_offset, column=1).font = Font(size=10, bold=True, name='微软雅黑', color="000000")
        r_offset += 1
    if column_title:
        ws.merge_cells(start_row=r_offset, start_column=matrix_c_offset, end_row=r_offset,
                       end_column=matrix_c_offset + matrix_column_num - 1)
        ws.cell(row=r_offset, column=matrix_c_offset).value = column_title
        ws.cell(row=r_offset, column=matrix_c_offset).alignment = alignment
        ws.cell(row=r_offset, column=matrix_c_offset).font = Font(size=10, bold=True, name='微软雅黑', color="000000")
        r_offset += 1
    if row_title:
        ws.merge_cells(start_row=matrix_r_offset, start_column=1, end_row=matrix_r_offset + matrix_row_num - 1,
                       end_column=1)
        ws.cell(row=matrix_r_offset, column=1).value = row_title
        ws.cell(row=matrix_r_offset, column=1).alignment = alignment
        ws.cell(row=matrix_r_offset, column=1).font = Font(size=10, bold=True, name='微软雅黑', color="000000")
        c_offset += 1
    # 标签
    if column_label_ls:
        for i in range(matrix_column_num):
            ws.cell(row=r_offset, column=matrix_c_offset + i).value = column_label_ls[i]
            ws.cell(row=r_offset, column=matrix_c_offset + i).alignment = alignment
            ws.cell(row=r_offset, column=matrix_c_offset + i).fill = PatternFill(patternType="solid",
                                                                                 start_color="33CCFF")
        r_offset += 1
    if row_label_ls:
        for i in range(matrix_row_num):
            ws.cell(row=matrix_r_offset + i, column=c_offset).value = row_label_ls[i]
            ws.cell(row=matrix_r_offset + i, column=c_offset).alignment = alignment
            ws.cell(row=matrix_r_offset + i, column=c_offset).fill = PatternFill(patternType="solid",
                                                                                 start_color="33CCFF")
        c_offset += 1
    # 校验，可省略
    if not (c_offset == matrix_c_offset and r_offset == matrix_r_offset):
        print("wrong here")
    for r_ in range(matrix_row_num):
        for c_ in range(matrix_column_num):
            ws.cell(row=matrix_r_offset + r_, column=matrix_c_offset + c_).value = matrix[r_][c_]

    # 先删除文件，再写入
    remove(file_path_abs)
    wb.save(file_path_abs)
    # print(file_name, 'wrote')


if __name__ == '__main__':
    mat = np.random.rand(4, 3)

    write_excel_with_matrix(os.path.join(os.getcwd(), 'test.xlsx'), "566666", mat, ['红', '绿', '蓝'],
                            ['1 1', '1 2', '2 1', '2 2'],
                            "这里是column_title 啊啊啊",
                            "这里是row_title 啊啊啊", "总标题 total")
    mat_2 = np.ones([4, 4])
    write_excel_with_matrix(os.path.join(os.getcwd(), 'test2.xlsx'), "result结构说明", mat_2, ['列', '向', '标', '签'],
                            ['行', '向', '标', '签'],
                            "转移目标（预测）",
                            "转移起点（已知）", "总标题")
